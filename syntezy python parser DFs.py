# -*- coding: utf-8 -*-
"""
Created on Sun Feb 11 21:28:42 2018

@author: Bartosz
"""
from datetime import datetime, date, timedelta
from openpyxl import Workbook
import openpyxl
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog

#os.getcwd()
#os.chdir('/home/bartosz/Desktop/Stepan/')


root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()


wb = openpyxl.load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames
dir(wb)
print(wb)
help(wb)

print(wb.sheetnames)

len('')

def number_export_from_string(string):
    tbl = ''
    try:
        if type(string) == str and len(string)>0:
            for i in string:
                if i.isdigit() == True:
                    tbl = tbl + i
            return int(tbl)
        elif type(string) == int:
            return string        
        else:
            return ''        
    except ValueError:
        pass
    



# Set the names of dataframe columns with material property values:

Table_columns = ["Synthesis number",
                 "ELN number",
                 "TSR number",
                 "Acid Value",
                 "Hydroxyl Value",
                 "Viscosity",
                 "Viscosity temperature (*C)",
                 "Water content(%)",
                 "Colour",
                 "Colour unit",
                 'Cycle time'
                 ]


# Set the dictionary of names of raws used for synthesis


raws_dictionary = {'Adipic acid':['AA','Adipic acid', 'adipic acid', 'Kwas adypinowy', 'kwas adypinowy', 'Kwas Adypinowy'],
                 'AGS':'AGS',
                 'Succinic acid':['SA succinic acid'],
                 'Phthalic Anhydride':['Phthalic Anhydride', 'Phthalic anhydride', 'PA', 'phthalic anhydride'],
                 'Isophthalic Acid':['Isophthalic Acid', 'isophthalic acid', 'Isophthalic acid', 'IPA', 'PIA'],
                 'Terephthalic acid':['Terephthalic Acid', 'terephthalic acid', 'Terephthalic acid', 'TPA', 'PTA'],
                 '1,2-Cyclohexane dicarboxylic acid anhydride':['1,2-Cyclohexane dicarboxylic acid anhydride', 'CHDA'],
                 'Maleic anhydride':['Maleic anhydride', 'maleic anhydride', 'MA'],
                 'Na-SIPA':['Na-SIPA', 'NA-SIPA', 'NASPIA', 'NaSIPA'],
                 'K-SIPA':['K-SIPA', 'KSIPA'],
                 'RA':['Ricinoleic Acid', 'RA', 'ricinoleic acid', 'Ricinoleic acid'],
                 'DRA':['DRA', 'dimeric ricinoleic acid'],
                 'MEG':['MEG', 'EG', 'Ethylene glycol', 'ethylene glycol', 'Monoethylene glycol'],
                 'DEG':['Diethylene glycol', 'DEG', 'diethylene glycol'],
                 'TEG':['TEG', 'triethylene glycol'],
                 'MP-diol':['MPdiol', 'MP-diol', 'MP diol', 'MPDiol'],
                 'NPG':['NPG,', 'Neopentyl glycol'],
                 'HPHP': ['HPHP'],
                 '1,2-PG':['PG', '1,2-PG', 'Propylene glycol', '1,2-Propylene Glycol', '1,2PG', '1,2-PD', '1,2-propanediol'],
                 'BDO':['1,4-BDO', 'BDO', '1,4 BDO', '1,4BDO', '1,4-butanediol'],
                 'HDO':['1,6-HDO', '1,6-hexanediol', '1,6HDO'],
                 'PEG400':['PEG400'],
                 'PEG200':['PEG200'],
                 'Carpol EDAP-800':['Carpol EDAP-800'],
                 'Resanon 100':['Resanon 100'],
                 'Polyol R3530':['Polyol R3530'],
                 'SBO': ['SBO', 'Soybean oil', 'Soybean Oil'],
                 'Castor Oil':['Castor oil', 'castor oil', 'CO'],
                 'TCPP':['TCPP'],
                 'Agent 601':['Agent', 'Agent 601'],
                 'TnBT': ['Tyzor TnBT', 'TnBT'],
                 'TPT':['TPT'],
                 'Tyzor LA':['Tyzor LA'],
                 'Condensate':['cond1', 'cond', 'Cond', 'condensate', 'cond.', 'H2O', 'h2o'],
                 'H3PO4 (85 %)':['H3PO4 (85%)'],
                 'H3PO4 (75 %)':['H3PO4 (75%)'],
                 'H3PO4':['H3PO4', 'Phoshporic acid', 'Phoshporic Acid'],
                 }


# get a list of raw names to be set for column names in the data frame:

Raw_names_list = list(raws_dictionary)

# last column names:

description_dictionary = {'Objective':['Objective'], 
                          'Conclusions':['Conclusion', 'Conclusions']
                          }

description_names_list = list(description_dictionary)


Main_data_frame_columns = Table_columns + Raw_names_list + description_names_list


#create the main data frame with all data:

syntheses = pd.DataFrame(columns=[Main_data_frame_columns])


# get the current sheet:



for temp_sheet in sheet_names:
    
    sheet = wb.get_sheet_by_name(temp_sheet)
    #create a local data frame for gathering data from one sheet:
    
    sheet_dataframe = pd.DataFrame(columns=[Main_data_frame_columns],)
    
    # input Synthesis number:
    sheet_dataframe.at[0,Main_data_frame_columns[0]] = sheet.cell(column=1, row=1).value
    
    # input ELN number
    if sheet.cell(column=2, row=1).value != None:
        sheet_dataframe.at[0, Main_data_frame_columns[1]] = sheet.cell(column=2, row=1).value[4:]
    
    
    
    # input TSR number
    
    
    for i in range(1,5):
        if sheet.cell(column=i, row=1).value != None:
            try:
                if sheet.cell(column=i, row=1).value[0:3] in ["USA", "EUR", "GEN"]:
                    sheet_dataframe.at[0, Main_data_frame_columns[2]] = sheet.cell(column=i, row=1).value
            except TypeError:
                pass
    
    # acid value search and input:      
    
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value == 'Av' :
                sheet_dataframe.at[0, Main_data_frame_columns[3]] = sheet.cell(column=3, row=i).value
    
    
    # hydroxyl value input
    
    
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value == 'OHv' :
                sheet_dataframe.at[0, Main_data_frame_columns[4]] = sheet.cell(column=3, row=i).value
    
    
    # viscosity input
                
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if str(sheet.cell(column=1, row=i).value) in ['Visc / mPa*s', 'Visc'] :
                viscosity_value = sheet.cell(column=3, row=i).value

                sheet_dataframe.at[0, Main_data_frame_columns[5]] = number_export_from_string(sheet.cell(column=3, row=i).value)
                break
    
    # viscosity temperature input
                
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value in ['Visc / mPa*s', 'Visc'] :
                sheet_dataframe.at[0, Main_data_frame_columns[6]] = number_export_from_string(sheet.cell(column=4, row=i).value)
                break
    
    # water content input
                
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value in ['H2O', 'h2o'] :
                sheet_dataframe.at[0, Main_data_frame_columns[7]] = sheet.cell(column=3, row=i).value
    
    
    # colour input
                
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value in ['colour', 'Colour', 'Color', 'color'] :
                sheet_dataframe.at[0, Main_data_frame_columns[8]] = sheet.cell(column=3, row=i).value
    
        # input color unit
    for i in range(15,35):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value in ['colour', 'Colour', 'Color', 'color'] :
                sheet_dataframe.at[0, Main_data_frame_columns[9]] = sheet.cell(column=4, row=i).value
    
    
    # get the cycle time
    for i in range(5,15):
        for j in range(5, 10):
            if sheet.cell(column=j, row=i).value != None and sheet.cell(column=j, row=i).value in ['time [h:min]']:
                t = datetime.strptime("00:00:00","%H:%M:%S")
                cycle_time = timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
                for t in range(1,100):
                    try:
                        if datetime.combine(date.min, sheet.cell(column=j, row=i+t+2).value) > datetime.combine(date.min, sheet.cell(column=j, row=i+t+1).value):
                            cycle_time = cycle_time + datetime.combine(date.min, sheet.cell(column=j, row=i+t+2).value) - datetime.combine(date.min, sheet.cell(column=j, row=i+t+1).value)
                    except TypeError:
                        pass
                sheet_dataframe.at[0, Main_data_frame_columns[10]] = cycle_time
                
    
    
    # this iterates thorugh the raws in the sheet data frame and searches the sheet raws for the match with dictionary.
    # When found it adds the amount of raw in a cell next to the 
    
    for raw in raws_dictionary:
        for i in range(5,20):
            if sheet.cell(column=1, row=i).value != None and str:
                if str(sheet.cell(column=1, row=i).value) in raws_dictionary[raw] :
                    sheet_dataframe.at[0, raw] = sheet.cell(column=2, row=i).value
    
    # insert objective
    for i in range(1,5):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value in ["Objective", "objective"]:
                sheet_dataframe.at[0, Main_data_frame_columns[48]] = sheet.cell(column=2, row=i).value
    
    # insert conclusions
    for i in range(1,5):
        if sheet.cell(column=1, row=i).value != None:
            if sheet.cell(column=1, row=i).value in ["Conclusion", "conclusion"]:
                sheet_dataframe.at[0, Main_data_frame_columns[49]] = sheet.cell(column=2, row=i).value
    
    syntheses = syntheses.append(sheet_dataframe, ignore_index=True)

# str.split(file_path, sep='/')[-1]

os.chdir('/'.join(str.split(file_path, sep='/')[:-1]))
#os.getcwd()
syntheses.to_csv('/'.join(str.split(file_path, sep='/')[:-1]) + '/' + str.split(file_path, sep='/')[-1] + '.csv',
                 sep='\t', encoding='utf-8', index=False, decimal=',')

